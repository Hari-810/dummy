import pandas as pd
import numpy as np
import spacy
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from sentence_transformers import SentenceTransformer
from tqdm import tqdm
from openpyxl import load_workbook

# Load NLP models
nlp = spacy.load("en_core_web_sm")
model = SentenceTransformer("all-MiniLM-L6-v2")

# File paths
input_excel = "AQM_TCG_Metrics_OriginalADO 1.xlsx"
output_excel = "AQM_TCG_Metrics_OriginalADO_Scored.xlsx"

# Column names
user_story_col = 'User Story Title'
ado_action_col = 'Unnamed: 6'
generated_action_col = 'Unnamed: 11'

# Read the Excel into a DataFrame (preserve as-is)
df = pd.read_excel(input_excel, dtype=str)
df[ado_action_col] = df[ado_action_col].fillna("").str.strip()
df[generated_action_col] = df[generated_action_col].fillna("").str.strip()
df[user_story_col] = df[user_story_col].fillna("").str.strip()

# Result storage for individual scores (needed to calculate averages)
tfidf_scores, contextual_precisions, contextual_recalls, contextual_f1s = [], [], [], []

# Grouped scoring by User Story Title
for story, group in tqdm(df.groupby(user_story_col), desc="Processing User Stories"):
    ado_actions = group[ado_action_col].tolist()
    generated_actions = group[generated_action_col].tolist()
    generated_actions = [x for x in generated_actions if x]

    if not generated_actions:
        tfidf_scores.extend([None] * len(group))
        contextual_precisions.extend([None] * len(group))
        contextual_recalls.extend([None] * len(group))
        contextual_f1s.extend([None] * len(group))
        continue

    gen_tfidf = TfidfVectorizer().fit(generated_actions)
    gen_vectors = gen_tfidf.transform(generated_actions)
    gen_embeddings = model.encode(generated_actions, convert_to_tensor=False)  # numpy array

    for ado in ado_actions:
        if not ado:
            tfidf_scores.append(None)
            contextual_precisions.append(None)
            contextual_recalls.append(None)
            contextual_f1s.append(None)
            continue

        # TFIDF similarity
        ado_vec = gen_tfidf.transform([ado])
        tfidf_sim = cosine_similarity(ado_vec, gen_vectors)[0]
        tfidf_best = np.max(tfidf_sim)

        # Contextual similarity embeddings
        ado_emb = model.encode(ado, convert_to_tensor=False).reshape(1, -1)  # shape (1, dim)

        # Precision: max cosine similarity of ADO to generated actions
        precision_vals = cosine_similarity(ado_emb, gen_embeddings)[0]
        precision = np.max(precision_vals) if len(precision_vals) > 0 else 0

        # Recall: average cosine similarity of generated actions to ADO (same direction cosine sim is symmetric)
        recall = np.mean(precision_vals) if len(precision_vals) > 0 else 0

        # F1 score harmonic mean of precision and recall
        if precision + recall > 0:
            f1 = 2 * precision * recall / (precision + recall)
        else:
            f1 = 0

        tfidf_scores.append(int(round(tfidf_best * 100)))
        contextual_precisions.append(int(round(precision * 100)))
        contextual_recalls.append(int(round(recall * 100)))
        contextual_f1s.append(int(round(f1 * 100)))

# Compute average scores per User Story group
grouped = df.groupby(user_story_col)
group_averages = {}

for story, group in grouped:
    indices = group.index.tolist()
    group_tfidf = [tfidf_scores[i] for i in indices if tfidf_scores[i] is not None]
    group_prec = [contextual_precisions[i] for i in indices if contextual_precisions[i] is not None]
    group_rec = [contextual_recalls[i] for i in indices if contextual_recalls[i] is not None]
    group_f1 = [contextual_f1s[i] for i in indices if contextual_f1s[i] is not None]

    avg_tfidf = int(round(np.mean(group_tfidf))) if group_tfidf else None
    avg_prec = int(round(np.mean(group_prec))) if group_prec else None
    avg_rec = int(round(np.mean(group_rec))) if group_rec else None
    avg_f1 = int(round(np.mean(group_f1))) if group_f1 else None

    group_averages[story] = (avg_tfidf, avg_prec, avg_rec, avg_f1)

# Load workbook without disturbing formatting
wb = load_workbook(input_excel)
ws = wb.active

# Determine where to put the new columns (after the last existing one)
start_col = ws.max_column + 1

# Write headers for average score columns at row 2
avg_headers = [
    "Avg TFIDF Score (%)",
    "Avg Contextual Precision (%)",
    "Avg Contextual Recall (%)",
    "Avg Contextual F1 Score (%)"
]
for i, header in enumerate(avg_headers):
    ws.cell(row=2, column=start_col + i, value=header)

# Write average scores only once per User Story group, on the first row of that group
row_offset = 3  # data starts at Excel row 3
for story, group in grouped:
    avg_tfidf, avg_prec, avg_rec, avg_f1 = group_averages[story]
    first_row = group.index.min() + row_offset  # Excel row number for first row of group

    ws.cell(row=first_row, column=start_col, value=avg_tfidf)
    ws.cell(row=first_row, column=start_col + 1, value=avg_prec)
    ws.cell(row=first_row, column=start_col + 2, value=avg_rec)
    ws.cell(row=first_row, column=start_col + 3, value=avg_f1)

# Save the workbook
wb.save(output_excel)

print(f"✅ Final Excel with grouped average scores saved to: {output_excel}")
